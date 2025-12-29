from datetime import datetime, timedelta
import pandas as pd
import numpy as np
from database.ms_sql_connection import fetch_query
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta
from environment.settings import config
from win32com.client import Dispatch


def run_euthanasia_report(month, year):
    # First day of the current month
    current_date = datetime(year, month, 1)

    # First day of the previous month
    prev_month_date = current_date - relativedelta(months=1)
    first_day_prev_month = prev_month_date.replace(day=1)

    query = f'''
    SELECT TOP (100) PERCENT 
        dbo.Animal.AnimalID, dbo.Animal.Name, 
        dbo.refSpecies.Species, dbo.AnimalDetails.DateOfBirth, 
        dbo.Animal.Sex, dbo.txnVisit.IntakeType, 
        dbo.txnVisit.tin_DateCreated AS IntakeDate, 
        dbo.Euthanasia.DateCreated AS EuthanizedDate, 
        DATEDIFF(day, dbo.txnVisit.tin_DateCreated, dbo.Euthanasia.DateCreated) AS TimetoEuth
    FROM dbo.Animal INNER JOIN
        dbo.AnimalDetails ON dbo.Animal.AnimalID = dbo.AnimalDetails.AnimalID INNER JOIN
        dbo.refSpecies ON dbo.Animal.SpeciesID = dbo.refSpecies.SpeciesID INNER JOIN
        dbo.Euthanasia ON dbo.Animal.AnimalID = dbo.Euthanasia.AnimalID INNER JOIN
        dbo.txnVisit ON dbo.Animal.AnimalID = dbo.txnVisit.AnimalID
    WHERE (dbo.refSpecies.Species IN ('Cat', 'Dog')) 
        AND (dbo.txnVisit.IntakeType IN ('TransferIn', 'OwnerSurrender', 'Stray', '[Return]')) 
        AND (DATEDIFF(day, dbo.txnVisit.tin_DateCreated, dbo.Euthanasia.DateCreated) >= 4
        AND DATEDIFF(day, dbo.txnVisit.tin_DateCreated, dbo.Euthanasia.DateCreated) <= 21) AND 
        dbo.Euthanasia.DateCreated>= '{first_day_prev_month}'
        AND dbo.Euthanasia.DateCreated < '{current_date}'
    ORDER BY dbo.Animal.AnimalID, IntakeDate, EuthanizedDate
    '''


    # In[4]:


    df1 = fetch_query(query)
    df1['IntakeDate'] = pd.to_datetime(df1['IntakeDate'])
    df1['EuthanizedDate'] = pd.to_datetime(df1['EuthanizedDate'])



    # # DATA CLEANING

    # In[5]:


    # keep only one intake per animal (closest to euth) 
    df1_refined=df1.groupby(['AnimalID']).agg({'TimetoEuth': pd.Series.min}).reset_index()
    df1_refined=pd.merge(df1_refined, df1, on=['AnimalID', 'TimetoEuth'], how='inner')


    # In[6]:


    # remove entries of intakes that occur on the same day

    # checks if animal id above is the same
    df1_refined['flag'] = np.where(df1_refined['AnimalID']==df1_refined['AnimalID'].shift(),1,0)
    # checks if intake date above is the same
    df1_refined['IntakeDate_dt'] = df1_refined['IntakeDate'].dt.date
    df1_refined['flag2'] = np.where(df1_refined['IntakeDate_dt']==df1_refined['IntakeDate_dt'].shift(),1,0)
    df1_refined['Intake_remove'] = np.where(df1_refined['flag']+df1_refined['flag2']==2,
                                                'REMOVE', 'KEEP')


    # In[7]:


    # delete rows based on Intake_remove column
    df1_refined.drop(df1_refined.loc[df1_refined['Intake_remove']=='REMOVE'].index, inplace=True)

    # delete unnecessary columns
    del df1_refined['flag']
    del df1_refined['flag2']
    del df1_refined['IntakeDate_dt']
    del df1_refined['Intake_remove']


    # In[8]:


    # reorder columns
    final = df1_refined[['AnimalID', 'Name', 'Species',  'DateOfBirth', 'Sex','IntakeType', 'IntakeDate', 'EuthanizedDate','TimetoEuth']]


   # File output
    output_file_name = f"{(current_date).strftime('%b')}-delayedeuthanasia.xlsx"
    output_path = f"{config.SERVER_PATH}/delayed_euthanasia/monthly/{output_file_name}"

    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        final.to_excel(writer, sheet_name='monthly', startrow=3, index=False)

    # Add headers and merge cells
    wb = load_workbook(output_path)
    ws = wb['monthly']
    ws['A1'] = f'Records for {current_date.strftime("%b-%Y")}'
    ws['A2'] = 'Intake to Euthanasia: 4-21 days inclusive'
    ws.merge_cells('A1:C1')
    ws.merge_cells('A2:D2')
    wb.save(output_path)

    # Auto-fit columns with Excel COM
    excel = Dispatch('Excel.Application')
    wb_com = excel.Workbooks.Open(output_path)
    for i in range(1, len(excel.Worksheets) + 1):
        excel.Worksheets(i).Activate()
        excel.ActiveSheet.Columns.AutoFit()
    wb_com.Close(True)


    return 
    




